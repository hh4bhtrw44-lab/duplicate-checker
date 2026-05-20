#!/usr/bin/env python3
"""客户查重管理系统 - Flask + SQLite (完整功能版)"""
import re
import os
import io
import csv
import hashlib
import sqlite3
import phonenumbers
from phonenumbers import geocoder, carrier
from datetime import datetime, timedelta, timezone

# 东八区时间 (北京时间)
BJ_TZ = timezone(timedelta(hours=8))
def bj_now():
    return datetime.now(BJ_TZ)
def bj_now_str():
    return bj_now().strftime("%Y-%m-%d %H:%M:%S")
def bj_date_str():
    return bj_now().strftime("%Y-%m-%d")
from functools import wraps
from flask import Flask, request, render_template, jsonify, session, redirect, url_for, g, send_file
import jieba
import jieba.analyse

app = Flask(__name__)

# ========= 云部署模式：不启动本地隧道 =========

# 稳定密钥
SECRET_KEY_FILE = os.path.join(app.root_path, '.secret_key')
if os.path.exists(SECRET_KEY_FILE):
    with open(SECRET_KEY_FILE, 'r') as f:
        app.secret_key = f.read().strip()
else:
    app.secret_key = os.urandom(24).hex()
    with open(SECRET_KEY_FILE, 'w') as f:
        f.write(app.secret_key)

# 持久化存储：优先使用 /app/data 目录（Railway Volume挂载点）
DATA_DIR = '/app/data'
if not os.path.exists(DATA_DIR):
    DATA_DIR = app.root_path
DB_PATH = os.path.join(DATA_DIR, 'data.db')

# 删除/导出/导入数据密码（不同于登录密码）
DATA_PASSWORD = 'sufahui520'

# 迁移旧数据

# 迁移旧数据：如果持久化目录有新的数据库且根目录有旧数据库则合并
OLD_DB = os.path.join(app.root_path, 'data.db')
if DATA_DIR != app.root_path and os.path.exists(OLD_DB) and not os.path.exists(DB_PATH):
    import shutil
    shutil.copy2(OLD_DB, DB_PATH)
    print(f"✅ 已迁移数据文件到 {DB_PATH}")

app.config['DATABASE'] = DB_PATH

# ========= 数据库初始化 =========

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(app.config['DATABASE'])
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
    return g.db

@app.teardown_appcontext
def close_db(exception):
    db = g.pop('db', None)
    if db is not None:
        db.close()

def init_db():
    conn = sqlite3.connect(app.config['DATABASE'])
    conn.execute("PRAGMA journal_mode=WAL")
    c = conn.cursor()
    # 为旧数据库兼容添加 phone_region 列
    try:
        c.execute("ALTER TABLE customers ADD COLUMN phone_region TEXT DEFAULT ''")
    except:
        pass
    c.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT DEFAULT 'admin',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT,
            email TEXT,
            company TEXT,
            notes TEXT,
            content TEXT DEFAULT '',
            phone_region TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS check_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer1_id INTEGER,
            customer2_id INTEGER,
            similarity REAL,
            level TEXT,
            checked_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (customer1_id) REFERENCES customers(id),
            FOREIGN KEY (customer2_id) REFERENCES customers(id)
        );

        CREATE TABLE IF NOT EXISTS check_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            check_date TEXT,
            check_type TEXT,
            duplicate_count INTEGER DEFAULT 0,
            total_checked INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    ''')
    try:
        c.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                  ('admin', hashlib.sha256('admin123'.encode()).hexdigest(), 'admin'))
    except sqlite3.IntegrityError:
        pass
    # add db indexes
    for idx_sql in [
        "CREATE INDEX IF NOT EXISTS idx_customers_name ON customers(name)",
        "CREATE INDEX IF NOT EXISTS idx_customers_phone ON customers(phone)",
        "CREATE INDEX IF NOT EXISTS idx_customers_created_at ON customers(created_at)",
        "CREATE INDEX IF NOT EXISTS idx_customers_phone_region ON customers(phone_region)",
        "CREATE INDEX IF NOT EXISTS idx_customers_company ON customers(company)",
    ]:
        try:
            c.execute(idx_sql)
        except:
            pass
    conn.commit()
    conn.close()

with app.app_context():
    init_db()

# ========= 登录装饰器 =========

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

# ========= 查重算法 =========

def preprocess_text(text):
    text = re.sub(r'[^\w\s]', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text.lower()

def split_words(text):
    words = jieba.lcut(text)
    return [w.strip() for w in words if w.strip() and len(w.strip()) > 1]

def get_shingles(words, k=3):
    if len(words) < k:
        return set([' '.join(words)])
    return set(' '.join(words[i:i+k]) for i in range(len(words) - k + 1))

def jaccard_similarity(set1, set2):
    if not set1 or not set2:
        return 0.0
    intersection = len(set1 & set2)
    union = len(set1 | set2)
    return intersection / union if union > 0 else 0.0

def simhash_value(words):
    hash_size = 64
    v = [0] * hash_size
    for word in words[:200]:
        h = int(hashlib.md5(word.encode('utf-8')).hexdigest(), 16)
        for i in range(hash_size):
            bit = (h >> i) & 1
            v[i] += 1 if bit else -1
    fingerprint = 0
    for i in range(hash_size):
        if v[i] > 0:
            fingerprint |= (1 << i)
    return fingerprint

def simhash_similarity(fp1, fp2):
    x = fp1 ^ fp2
    return 1.0 - (bin(x).count('1') / 64.0)

def check_duplicate(text1, text2):
    clean1 = preprocess_text(text1)
    clean2 = preprocess_text(text2)
    words1 = split_words(clean1)
    words2 = split_words(clean2)

    if not words1 or not words2:
        return {"jaccard": 0, "simhash": 0, "overall": 0, "level": "文本过短", "word_count_1": len(words1), "word_count_2": len(words2)}

    shingles1 = get_shingles(words1, 3)
    shingles2 = get_shingles(words2, 3)
    jaccard = jaccard_similarity(shingles1, shingles2)

    fp1 = simhash_value(words1)
    fp2 = simhash_value(words2)
    simhash = simhash_similarity(fp1, fp2)

    overall = jaccard * 0.6 + simhash * 0.4

    if overall >= 0.85:
        level = "🔴 高度重复"
    elif overall >= 0.60:
        level = "🟡 中度重复"
    elif overall >= 0.30:
        level = "🟠 轻度重复"
    else:
        level = "🟢 基本不重复"

    return {
        "jaccard": round(jaccard * 100, 2),
        "simhash": round(simhash * 100, 2),
        "overall": round(overall * 100, 2),
        "level": level,
        "word_count_1": len(words1),
        "word_count_2": len(words2)
    }

# ========= 路由：认证 =========

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.get_json() or request.form
        username = data.get('username', '')
        password = data.get('password', '')
        hashed = hashlib.sha256(password.encode()).hexdigest()
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE username=? AND password=?", (username, hashed)).fetchone()
        if user:
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            return jsonify({"ok": True})
        return jsonify({"ok": False, "error": "用户名或密码错误"}), 401
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/api/change-password', methods=['POST'])
@login_required
def api_change_password():
    data = request.get_json()
    old_pw = data.get('old_password', '')
    new_pw = data.get('new_password', '')

    if not old_pw or not new_pw:
        return jsonify({"error": "旧密码和新密码不能为空"}), 400
    if len(new_pw) < 6:
        return jsonify({"error": "新密码长度不能少于6位"}), 400

    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (session['user_id'],)).fetchone()
    if not user:
        return jsonify({"error": "用户不存在"}), 404

    if user['password'] != hashlib.sha256(old_pw.encode()).hexdigest():
        return jsonify({"error": "旧密码错误"}), 401

    db.execute("UPDATE users SET password=? WHERE id=?",
               (hashlib.sha256(new_pw.encode()).hexdigest(), session['user_id']))
    db.commit()
    return jsonify({"ok": True, "message": "密码修改成功"})

# ========= 路由：管理界面 =========

@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/lite')
@login_required
def index_lite():
    return render_template('lite.html')

@app.route('/api/customers', methods=['GET'])
@login_required
def api_customers():
    db = get_db()
    search = request.args.get('search', '').strip()
    search_field = request.args.get('field', 'all')
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))
    offset = (page - 1) * per_page

    region_filter = request.args.get('region', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    company_filter = request.args.get('company', '').strip()

    if per_page > 100:
        per_page = 100

    where_clauses = []
    params = []

    if search:
        if search_field == 'name':
            where_clauses.append("name LIKE ?")
            params.append(f'%{search}%')
        elif search_field == 'phone':
            where_clauses.append("phone LIKE ?")
            params.append(f'%{search}%')
        elif search_field == 'email':
            where_clauses.append("email LIKE ?")
            params.append(f'%{search}%')
        elif search_field == 'company':
            where_clauses.append("company LIKE ?")
            params.append(f'%{search}%')
        else:
            where_clauses.append("(name LIKE ? OR phone LIKE ? OR email LIKE ? OR company LIKE ? OR notes LIKE ?)")
            params.extend([f'%{search}%'] * 5)

    if region_filter:
        where_clauses.append("phone_region = ?")
        params.append(region_filter)
    if date_from:
        where_clauses.append("DATE(created_at) >= ?")
        params.append(date_from)
    if date_to:
        where_clauses.append("DATE(created_at) <= ?")
        params.append(date_to)
    if company_filter:
        where_clauses.append("company LIKE ?")
        params.append(f'%{company_filter}%')

    where_sql = ""
    if where_clauses:
        where_sql = "WHERE " + " AND ".join(where_clauses)

    rows = db.execute(f"SELECT * FROM customers {where_sql} ORDER BY id DESC LIMIT ? OFFSET ?",
                      params + [per_page, offset]).fetchall()
    total = db.execute(f"SELECT COUNT(*) FROM customers {where_sql}", params).fetchone()[0]

    return jsonify({
        "customers": [dict(r) for r in rows],
        "total": total,
        "page": page,
        "per_page": per_page
    })

def clean_phone(phone):
    """清洗电话号码：去空格、去+号、去横线、去括号"""
    if not phone:
        return phone
    return re.sub(r'[\s\+\-\(\)]', '', phone)

def detect_phone_region(phone):
    """检测号码归属地，返回中文描述"""
    if not phone:
        return ''
    try:
        clean = phone.replace('+', '').replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
        # 先尝试解析带+号的完整号码
        try:
            parsed = phonenumbers.parse('+' + clean, None)
            desc = geocoder.description_for_number(parsed, 'zh')
            if desc:
                return desc
        except:
            pass
        # 中国大陆手机号 (11位, 1开头)
        if len(clean) == 11 and clean.startswith('1'):
            try:
                parsed = phonenumbers.parse(clean, "CN")
                desc = geocoder.description_for_number(parsed, 'zh')
                if desc:
                    return desc
            except:
                pass
        # 香港 8位
        if len(clean) == 8:
            try:
                parsed = phonenumbers.parse(clean, "HK")
                desc = geocoder.description_for_number(parsed, 'zh')
                if desc:
                    return desc
            except:
                pass
        # 台湾 9位 9开头
        if len(clean) == 9 and clean.startswith('9'):
            try:
                parsed = phonenumbers.parse(clean, "TW")
                desc = geocoder.description_for_number(parsed, 'zh')
                if desc:
                    return desc
            except:
                pass
        # 美国/加拿大 10位
        if len(clean) == 10:
            try:
                parsed = phonenumbers.parse(clean, "US")
                desc = geocoder.description_for_number(parsed, 'zh')
                if desc:
                    return desc
            except:
                pass
        # 最后尝试作为中国号码
        try:
            parsed = phonenumbers.parse(clean, "CN")
            country = geocoder.country_name_for_number(parsed, 'zh')
            if country:
                return country
        except:
            pass
        return ''
    except:
        return ''
@app.route('/api/customers/fix-data', methods=['POST'])
@login_required
def api_fix_customer_data():
    """自动修复数据：将电话栏中的姓名+电话混合数据拆分"""
    db = get_db()
    rows = db.execute('SELECT id, name, phone FROM customers').fetchall()
    fixed = 0
    import re as re_mod
    for r in rows:
        cid = r['id']
        name = (r['name'] or '').strip()
        phone = (r['phone'] or '').strip()

        # 情况1：电话栏里混了姓名
        if phone:
            parts = re_mod.split(r'[\s\t,，|;；]+', phone)
            phone_parts = []
            name_parts = []
            for p in parts:
                p = p.strip()
                if not p:
                    continue
                if re_mod.match(r'^[\+\d][\d\s\-\(\)]{4,}$', p):
                    phone_parts.append(p)
                else:
                    name_parts.append(p)

            if phone_parts and name_parts:
                new_phone = clean_phone(' '.join(phone_parts))
                new_name = ' '.join(name_parts)
                if name:
                    new_name = name + ' ' + new_name
                db.execute('UPDATE customers SET name=?, phone=? WHERE id=?', (new_name.strip(), new_phone, cid))
                fixed += 1
            elif not phone_parts and name_parts and not name:
                db.execute('UPDATE customers SET name=?, phone=? WHERE id=?', (name_parts[0], '', cid))
                fixed += 1

        # 情况2：姓名栏里混了电话
        if name:
            parts = re_mod.split(r'[\s\t,，|;；]+', name)
            phone_parts = []
            name_parts = []
            for p in parts:
                p = p.strip()
                if not p:
                    continue
                if re_mod.match(r'^[\+\d][\d\s\-\(\)]{4,}$', p):
                    phone_parts.append(p)
                else:
                    name_parts.append(p)

            if phone_parts and name_parts:
                new_name = ' '.join(name_parts)
                existing_phone = db.execute('SELECT phone FROM customers WHERE id=?', (cid,)).fetchone()['phone'] or ''
                new_phone = clean_phone((existing_phone + ' ' + ' '.join(phone_parts)).strip())
                if existing_phone and existing_phone == new_phone:
                    continue
                db.execute('UPDATE customers SET name=?, phone=? WHERE id=?', (new_name.strip(), new_phone, cid))
                fixed += 1

    db.commit()
    return jsonify({'ok': True, 'fixed': fixed, 'message': f'已修复 {fixed} 条数据'})

@app.route('/api/customers/quick-add', methods=['POST'])
@login_required
def api_quick_add():
    """快速添加：自动识别电话和姓名，返回去重结果"""
    data = request.get_json()
    lines = data.get('lines', '')
    if not lines:
        return jsonify({'error': '请输入客户信息'}), 400

    lines = [l.strip() for l in lines.split('\n') if l.strip()]
    db = get_db()
    imported = 0
    all_duplicates = []

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # 先处理整行：如果有+号但被空格/符号分割，先把所有数字段合并
        raw_digits = re.sub(r'[\+\-\s\(\)\.\/]', '', line)
        # 如果整行去掉符号后全是数字，直接当作号码处理
        if raw_digits.isdigit() and len(raw_digits) >= 5:
            phone = clean_phone(raw_digits)
            name = ''
        else:
            all_segments = re.split(r'[\s\t,，|;；]+', line)
            phone_candidates = []
            name_candidates = []
            prefix_buf = ''  # 存+86等短前缀
            digit_segments = []  # 收集连续数字段

            def flush_digits():
                nonlocal prefix_buf
                if digit_segments:
                    combined = ''.join(digit_segments)
                    digit_segments.clear()
                    if len(combined) >= 5:
                        phone_candidates.append(prefix_buf + combined)
                        prefix_buf = ''
                    else:
                        name_candidates.append(combined)

            for seg in all_segments:
                seg = seg.strip()
                if not seg:
                    continue
                # 判断是否是国际区号前缀（+86、+852等）
                if re.match(r'^\+', seg) and len(seg) <= 5:
                    flush_digits()
                    prefix_buf = re.sub(r'[\+\s]', '', seg)
                    continue
                digit_only = re.sub(r'[\+\-\s\(\)\.\/]', '', seg)
                if seg == '+' or (digit_only.isdigit() and seg.startswith('+')):
                    flush_digits()
                    # 去掉+号后的数字
                    prefix_buf = digit_only
                elif digit_only.isdigit() and len(digit_only) < 11:
                    # 短数字段合并起来
                    digit_segments.append(digit_only)
                elif digit_only.isdigit():
                    flush_digits()
                    phone_candidates.append(prefix_buf + digit_only)
                    prefix_buf = ''
                elif seg.isdigit():
                    digit_segments.append(seg)
                else:
                    flush_digits()
                    name_candidates.append(seg)
            flush_digits()

            # 如果前缀还留着，说明没有电话跟它，当姓名处理
            if prefix_buf and not phone_candidates:
                name_candidates.insert(0, '+' + prefix_buf)

            if not phone_candidates:
                combined = ''.join(name_candidates)
                found_nums = re.findall(r'\d{5,}', combined)
                if found_nums:
                    phone_candidates = found_nums
                    for n in found_nums:
                        combined = combined.replace(n, '', 1)
                    name_candidates = [combined.strip()] if combined.strip() else []

            name = ' '.join(name_candidates) if name_candidates else ''
            phone = clean_phone(' '.join(phone_candidates)) if phone_candidates else ''


        if not name and not phone:
            continue

        # 去重检查
        dup_entry = {'line': line, 'name': name, 'phone': phone, 'duplicates': []}
        if name:
            same = db.execute('SELECT id, name, phone, company, created_at FROM customers WHERE name=?', (name,)).fetchall()
            for s in same:
                dup_entry['duplicates'].append({
                    'id': s['id'], 'name': s['name'], 'phone': s['phone'],
                    'company': s['company'], 'field': '姓名', 'created_at': s['created_at']
                })
        if phone:
            same = db.execute('SELECT id, name, phone, company, created_at FROM customers WHERE phone=? AND phone!=""', (phone,)).fetchall()
            for s in same:
                if not any(d['id'] == s['id'] for d in dup_entry['duplicates']):
                    dup_entry['duplicates'].append({
                        'id': s['id'], 'name': s['name'], 'phone': s['phone'],
                        'company': s['company'], 'field': '电话', 'created_at': s['created_at']
                    })

        if dup_entry['duplicates']:
            all_duplicates.append(dup_entry)
        else:
            try:
                db.execute('INSERT INTO customers (name, phone, company, phone_region) VALUES (?, ?, ?, ?)',
                           (name, phone, '', ''))
                imported += 1
            except:
                pass

    db.commit()
    return jsonify({
        'ok': True, 'imported': imported, 'total': len(lines),
        'duplicates': all_duplicates
    })

@app.route('/api/customers/detect-regions', methods=['POST'])
@login_required
def api_detect_regions():
    """批量检测电话归属地"""
    db = get_db()
    rows = db.execute("SELECT id, phone FROM customers WHERE phone IS NOT NULL AND phone != ''").fetchall()
    updated = 0
    for r in rows:
        phone = r['phone']
        region = ''
        try:
            parsed = phonenumbers.parse('+' + phone.replace('+', ''), None)
            region = geocoder.description_for_number(parsed, 'zh') or geocoder.country_name_for_number(parsed, 'zh') or ''
        except:
            try:
                digits_only = re.sub(r"[^0-9]", "", phone)
                if len(digits_only) >= 7:
                    for cc_ in ["UZ", "KG", "KZ", "US", "GB", "RU"]:
                        try:
                            p2 = phonenumbers.parse("+" + digits_only, cc_)
                            if phonenumbers.is_valid_number(p2):
                                parsed = p2
                                region = geocoder.description_for_number(parsed, "zh") or geocoder.country_name_for_number(parsed, "zh") or ""
                                break
                        except:
                            pass
            except:
                pass
        if region:
            db.execute('UPDATE customers SET phone_region=? WHERE id=?', (region, r['id']))
            updated += 1
    db.commit()
    return jsonify({'ok': True, 'updated': updated, 'message': f'已检测 {updated} 个号码的归属地'})

@app.route('/api/customers', methods=['POST'])
@login_required
def api_add_customer():
    data = request.get_json()
    name = data.get('name', '').strip()
    phone = clean_phone(data.get('phone', ''))
    email = data.get('email', '').strip()
    company = data.get('company', '').strip()
    notes = data.get('notes', '').strip()
    content = data.get('content', '').strip()

    if not name and not phone:
        return jsonify({"error": "姓名和电话至少填一项"}), 400

    db = get_db()

    # ====== 去重检测 ======
    duplicates = []
    if name:
        same_name = db.execute("SELECT id, name, phone, company, created_at FROM customers WHERE name = ? AND id != ?",
                               (name, request.json.get('id', 0) or 0)).fetchall()
        for r in same_name:
            duplicates.append({"id": r["id"], "name": r["name"], "phone": r["phone"], "company": r["company"], "field": "姓名", "created_at": r["created_at"]})
    if phone:
        same_phone = db.execute("SELECT id, name, phone, company, created_at FROM customers WHERE phone = ? AND phone != '' AND id != ?",
                                (phone, request.json.get('id', 0) or 0)).fetchall()
        for r in same_phone:
            duplicates.append({"id": r["id"], "name": r["name"], "phone": r["phone"], "company": r["company"], "field": "电话", "created_at": r["created_at"]})

    if duplicates:
        # 去重 2：如果勾选了 force_add，强制添加
        if request.json.get('force_add', False):
            pass
        else:
            return jsonify({
                "duplicate_warning": True,
                "duplicates": duplicates,
                "message": f"发现 {len(duplicates)} 个可能重复的客户"
            }), 409

    # 检测号码归属地
    region = ''
    if phone:
        try:
            parsed = phonenumbers.parse('+' + phone.replace('+', ''), None)
            region = geocoder.description_for_number(parsed, 'zh') or geocoder.country_name_for_number(parsed, 'zh') or ''
        except:
            pass

    c = db.execute("INSERT INTO customers (name, phone, email, company, notes, content, phone_region) VALUES (?, ?, ?, ?, ?, ?, ?)",
                   (name, phone, email, company, notes, content, region))
    db.commit()
    return jsonify({"ok": True, "id": c.lastrowid})

@app.route('/api/customers/<int:id>', methods=['PUT'])
@login_required
def api_update_customer(id):
    data = request.get_json()
    name = data.get('name', '').strip()
    phone = clean_phone(data.get('phone', ''))
    email = data.get('email', '').strip()
    company = data.get('company', '').strip()
    notes = data.get('notes', '').strip()
    content = data.get('content', '').strip()

    if not name:
        return jsonify({"error": "客户姓名不能为空"}), 400

    db = get_db()

    # 编辑时去重检测
    duplicates = []
    if name:
        same_name = db.execute("SELECT id, name, phone, company, created_at FROM customers WHERE name = ? AND id != ?",
                               (name, id)).fetchall()
        for r in same_name:
            duplicates.append({"id": r["id"], "name": r["name"], "phone": r["phone"], "company": r["company"], "field": "姓名", "created_at": r["created_at"]})
    if phone:
        same_phone = db.execute("SELECT id, name, phone, company, created_at FROM customers WHERE phone = ? AND phone != '' AND id != ?",
                                (phone, id)).fetchall()
        for r in same_phone:
            duplicates.append({"id": r["id"], "name": r["name"], "phone": r["phone"], "company": r["company"], "field": "电话", "created_at": r["created_at"]})

    if duplicates:
        return jsonify({
            "duplicate_warning": True,
            "duplicates": duplicates,
            "message": f"检测到 {len(duplicates)} 个可能重复的客户"
        }), 409

    # 检测归属地
    region = ''
    if phone:
        try:
            parsed = phonenumbers.parse('+' + phone.replace('+', ''), None)
            region = geocoder.description_for_number(parsed, 'zh') or geocoder.country_name_for_number(parsed, 'zh') or ''
        except:
            pass
    db.execute("UPDATE customers SET name=?, phone=?, email=?, company=?, notes=?, content=?, phone_region=?, updated_at=bj_now_str() WHERE id=?",
               (name, phone, email, company, notes, content, region, id))
    db.commit()
    return jsonify({"ok": True})

@app.route('/api/customers/<int:id>', methods=['DELETE'])
@login_required
def api_delete_customer(id):
    pwd = request.args.get('pwd', '') or (request.get_json(silent=True) or {}).get('pwd', '')
    if pwd != DATA_PASSWORD:
        return jsonify({"error": "密码错误，无法删除"}), 403
    db = get_db()
    db.execute("DELETE FROM customers WHERE id=?", (id,))
    db.execute("DELETE FROM check_history WHERE customer1_id=? OR customer2_id=?", (id, id))
    db.commit()
    return jsonify({"ok": True})

@app.route('/api/customers/batch-delete', methods=['POST'])
@login_required
def api_batch_delete():
    data = request.get_json()
    ids = data.get('ids', [])
    pwd = data.get('pwd', '')
    if pwd != DATA_PASSWORD:
        return jsonify({"error": "密码错误，无法批量删除"}), 403
    if not ids:
        return jsonify({"error": "请选择要删除的客户"}), 400
    db = get_db()
    placeholders = ','.join('?' * len(ids))
    db.execute(f"DELETE FROM customers WHERE id IN ({placeholders})", ids)
    db.commit()
    return jsonify({"ok": True, "deleted": len(ids)})

# ========= 验证数据操作密码 =========
@app.route('/api/verify-data-password', methods=['POST'])
@login_required
def api_verify_data_password():
    data = request.get_json()
    pwd = data.get('pwd', '')
    if pwd == DATA_PASSWORD:
        return jsonify({'ok': True})
    return jsonify({'ok': False}), 403

# ========= 导出 Excel =========

@app.route('/api/export/excel', methods=['GET'])
@login_required
def api_export_excel():
    pwd = request.args.get('pwd', '')
    if pwd != DATA_PASSWORD:
        return jsonify({"error": "密码错误，无法导出数据"}), 403
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({"error": "导出功能需要安装 openpyxl：pip install openpyxl"}), 500

    db = get_db()

    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    region = request.args.get('region', '').strip()
    search = request.args.get('search', '').strip()

    where_clauses = []
    params = []
    if start_date:
        where_clauses.append("DATE(created_at) >= ?")
        params.append(start_date)
    if end_date:
        where_clauses.append("DATE(created_at) <= ?")
        params.append(end_date)
    if region:
        where_clauses.append("phone_region = ?")
        params.append(region)
    if search:
        where_clauses.append("(name LIKE ? OR phone LIKE ? OR company LIKE ? OR email LIKE ?)")
        params.extend([f'%{search}%'] * 4)

    where_sql = ""
    if where_clauses:
        where_sql = "WHERE " + " AND ".join(where_clauses)

    customers = db.execute(f"SELECT * FROM customers {where_sql} ORDER BY id DESC", params).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "客户数据"

    # 表头样式
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='667EEA', end_color='764BA2', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ['编号', '姓名', '电话', '邮箱', '公司', '备注', '查重内容', '创建时间', '更新时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 数据行
    cell_font = Font(name='微软雅黑', size=10)
    for row_idx, c in enumerate(customers, 2):
        values = [c['id'], c['name'], c['phone'], c['email'], c['company'],
                  c['notes'], c['content'], c['created_at'], c['updated_at']]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.alignment = Alignment(vertical='center')
            cell.border = thin_border

    # 列宽
    widths = [8, 16, 18, 24, 20, 24, 40, 20, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 写入内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = bj_now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'客户数据_{timestamp}.xlsx'
    )

# ========= 批量导入 =========

@app.route('/api/customers/import', methods=['POST'])
@login_required
def api_import_customers():
    if 'file' not in request.files:
        return jsonify({"error": "请上传文件"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "文件不能为空"}), 400

    # 解析文件
    records = []
    filename = file.filename.lower()

    try:
        if filename.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            for row in reader:
                records.append(row)
        elif filename.endswith('.xlsx'):
            try:
                import openpyxl
            except ImportError:
                return jsonify({"error": "导入xlsx需要安装openpyxl"}), 500
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            header_row = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                record = {}
                for i, val in enumerate(row):
                    if i < len(header_row) and header_row[i]:
                        record[header_row[i]] = str(val) if val is not None else ''
                if record:
                    records.append(record)
        elif filename.endswith('.txt'):
            content = file.read().decode('utf-8-sig')
            lines = [line.strip() for line in content.split('\n') if line.strip()]
            # 尝试识别格式
            for line in lines:
                record = {}
                parts = None
                
                # 尝试按常见分隔符分割
                for sep in ['\t', ',', '|', '，', ';', '；']:
                    test_parts = line.split(sep)
                    if len(test_parts) >= 2:
                        parts = test_parts
                        break
                
                # 如果常见分隔符都不行，尝试按空白分割（2个以上连续空格）
                if parts is None:
                    import re as re2
                    test_parts = re2.split(r'\s{2,}', line)  # 2个以上空格
                    if len(test_parts) >= 2:
                        parts = test_parts
                
                # 最后尝试单空格（但要排除电话号码中间的空格）
                if parts is None:
                    test_parts = line.split()
                    if len(test_parts) >= 2:
                        # 判断第一个是不是电话号码格式（包含+号或纯数字）
                        first = test_parts[0].strip()
                        if re.match(r'^[\+\d][\d\s\-\(\)]{4,}$', first.replace(' ', '')):
                            parts = test_parts
                        else:
                            # 可能有电话在第一列的情况，但也要考虑空格分割
                            for i, p in enumerate(test_parts):
                                p = p.strip()
                                if re.match(r'^[\+\d][\d\s\-\(\)]{6,}$', p.replace(' ', '')) and i > 0:
                                    # 在电话位置分割
                                    parts = [' '.join(test_parts[:i]), p] + test_parts[i+1:]
                                    break
                            if parts is None:
                                parts = [line]  # 无法智能分割，整行当姓名

                if parts is None:
                    parts = [line]

                if len(parts) >= 2:
                    first = parts[0].strip()
                    second = parts[1].strip()
                    if re.match(r'^[\+\d][\d\s\-\(\)]{4,}$', first):
                        record['电话'] = first
                        record['姓名'] = second
                    else:
                        record['姓名'] = first
                        record['电话'] = second
                    if len(parts) >= 3:
                        record['公司'] = parts[2].strip()
                    if len(parts) >= 4:
                        record['备注'] = parts[3].strip()
                    records.append(record)
                elif len(parts) == 1 and line:
                    record['姓名'] = parts[0].strip()
                    records.append(record)
        else:
            return jsonify({"error": "仅支持 CSV、XLSX 和 TXT 格式"}), 400
    except Exception as e:
        return jsonify({"error": f"文件解析失败: {str(e)}"}), 400

    if not records:
        return jsonify({"error": "文件中没有数据"}), 400

    # 字段映射
    field_map = {
        '姓名': 'name', '名字': 'name', 'name': 'name',
        '电话': 'phone', '手机': 'phone', 'phone': 'phone', 'tel': 'phone',
        '邮箱': 'email', '邮件': 'email', 'email': 'email',
        '公司': 'company', '企业': 'company', '单位': 'company', 'company': 'company',
        '备注': 'notes', '备注': 'notes', 'notes': 'notes',
        '查重内容': 'content', '内容': 'content', 'content': 'content',
    }

    db = get_db()
    imported = 0
    errors = []

    for idx, record in enumerate(records):
        mapped = {}
        for k, v in record.items():
            key = k.strip()
            if key in field_map:
                mapped[field_map[key]] = v.strip() if v else ''

        # 智能查找字段
        if 'name' not in mapped:
            for k in record:
                if any(term in k for term in ['姓名', '名字', '名称', 'name']):
                    mapped['name'] = record[k].strip() if record[k] else ''
                    break

        name = mapped.get('name', '')
        phone = clean_phone(mapped.get('phone', ''))
        email = mapped.get('email', '')

        if not name and not phone:
            errors.append(f"第{idx+2}行：姓名和电话均为空，跳过")
            continue

        try:
            db.execute("INSERT INTO customers (name, phone, email, company, notes, content) VALUES (?, ?, ?, ?, ?, ?)",
                       (name, phone, email, mapped.get('company', ''), mapped.get('notes', ''), mapped.get('content', '')))
            imported += 1
        except Exception as e:
            errors.append(f"第{idx+2}行（{name}）：导入失败 - {str(e)}")

    db.commit()

    result = {
        "ok": True,
        "imported": imported,
        "total": len(records),
        "errors": errors[:20]  # 最多返回20条错误
    }
    if errors:
        result["has_errors"] = True
        if len(errors) > 20:
            result["error_summary"] = f"共{len(errors)}条错误，仅显示前20条"

    return jsonify(result)

# ========= 电话号码归属地检测 =========

@app.route('/api/phone-lookup', methods=['POST'])
@login_required
def api_phone_lookup():
    data = request.get_json()
    phone = data.get('phone', '').strip()
    if not phone:
        return jsonify({"error": "请输入电话号码"}), 400

    clean_phone = re.sub(r'[\s\-\(\)]', '', phone)
    
    try:
        # Try multiple formats
        parsed = None
        results = []
        
        if not clean_phone.startswith('+'):
            clean_phone_with_plus = '+' + clean_phone
        else:
            clean_phone_with_plus = clean_phone

        # Try to detect - try various countries
        for test in [clean_phone_with_plus, clean_phone]:
            for cc in [None, "CN", "US", "GB", "JP", "KR", "SG", "TW", "HK", "TH", "MY", "VN", "PH", "ID", "AU", "CA", "DE", "FR", "IT", "ES", "RU", "IN", "UZ", "KG", "KZ", "UZ", "KG", "KZ"]:
                try:
                    p = phonenumbers.parse(test, cc)
                    if phonenumbers.is_valid_number(p):
                        results.append(p)
                        break
                except:
                    pass
            if results:
                break

        if not results:
            if clean_phone.startswith('86') and len(clean_phone) > 7:
                try:
                    p = phonenumbers.parse('+' + clean_phone, "CN")
                    if phonenumbers.is_valid_number(p):
                        results.append(p)
                except:
                    pass
            if not results:
                try:
                    p = phonenumbers.parse(clean_phone, None)
                    results.append(p)
                except:
                    pass

        if not results:
            return jsonify({"phone": phone, "valid": False, "country": "未知", "country_code": ""})

        parsed = results[0]

        if not phonenumbers.is_valid_number(parsed):
            return jsonify({"phone": phone, "valid": False, "country": "未知", "country_code": ""})

        # Get info
        country = geocoder.country_name_for_number(parsed, "zh")
        if not country:
            country = geocoder.country_name_for_number(parsed, "en")
        if not country:
            country = geocoder.description_for_number(parsed, "zh")
        if not country:
            country = "未知"
        region = geocoder.description_for_number(parsed, "zh")
        carrier_name = carrier.name_for_number(parsed, "zh")
        country_code = str(parsed.country_code)

        return jsonify({
            "phone": phone, "valid": True,
            "country": country, "region": region or country,
            "carrier": carrier_name or "",
            "country_code": "+" + country_code,
            "international": phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.INTERNATIONAL)
        })
    except Exception as e:
        return jsonify({"phone": phone, "valid": False, "country": "未知", "country_code": "", "error": str(e)})

# ========= 统计与图表分析 =========

@app.route('/api/analytics/overview')
@login_required
def api_analytics_overview():
    db = get_db()

    # 客户总数
    total_customers = db.execute("SELECT COUNT(*) FROM customers").fetchone()[0]

    # 今日新增
    today = bj_now().strftime("%Y-%m-%d")
    today_new = db.execute("SELECT COUNT(*) FROM customers WHERE DATE(created_at) = ?", (today,)).fetchone()[0]

    # 本周新增
    week_ago = (bj_now() - timedelta(days=7)).strftime("%Y-%m-%d")
    week_new = db.execute("SELECT COUNT(*) FROM customers WHERE DATE(created_at) >= ?", (week_ago,)).fetchone()[0]

    # 本月新增
    month_start = bj_now().strftime("%Y-%m-01")
    month_new = db.execute("SELECT COUNT(*) FROM customers WHERE DATE(created_at) >= ?", (month_start,)).fetchone()[0]

    # 今日查重
    today_checks = db.execute("SELECT COUNT(*) FROM check_history WHERE DATE(checked_at) = ?", (today,)).fetchone()[0]

    # 总查重次数
    total_checks = db.execute("SELECT COUNT(*) FROM check_history").fetchone()[0]

    # 高度重复记录数（>85%）
    high_duplicate = db.execute("SELECT COUNT(*) FROM check_history WHERE similarity >= 85").fetchone()[0]

    # 中度重复（60-85%）
    mid_duplicate = db.execute("SELECT COUNT(*) FROM check_history WHERE similarity >= 60 AND similarity < 85").fetchone()[0]

    # 轻度重复（30-60%）
    low_duplicate = db.execute("SELECT COUNT(*) FROM check_history WHERE similarity >= 30 AND similarity < 60").fetchone()[0]

    # 最近30天每日新增趋势
    daily_new = []
    for i in range(29, -1, -1):
        day = (bj_now() - timedelta(days=i)).strftime("%Y-%m-%d")
        count = db.execute("SELECT COUNT(*) FROM customers WHERE DATE(created_at) = ?", (day,)).fetchone()[0]
        daily_new.append({"date": day, "count": count})

    # 最近30天查重趋势
    daily_checks = []
    for i in range(29, -1, -1):
        day = (bj_now() - timedelta(days=i)).strftime("%Y-%m-%d")
        count = db.execute("SELECT COUNT(*) FROM check_history WHERE DATE(checked_at) = ?", (day,)).fetchone()[0]
        daily_checks.append({"date": day, "count": count})

    # 相似度等级分布
    distribution = {
        "high": high_duplicate,
        "mid": mid_duplicate,
        "low": low_duplicate,
        "none": total_checks - high_duplicate - mid_duplicate - low_duplicate
    }

    # 查重次数最多的10个客户
    top_checked = db.execute("""
        SELECT c.id, c.name, c.phone, COUNT(*) as check_count,
               ROUND(AVG(h.similarity), 2) as avg_similarity
        FROM check_history h
        JOIN customers c ON h.customer1_id = c.id
        GROUP BY c.id
        ORDER BY check_count DESC
        LIMIT 10
    """).fetchall()

    return jsonify({
        "total_customers": total_customers,
        "today_new": today_new,
        "week_new": week_new,
        "month_new": month_new,
        "today_checks": today_checks,
        "total_checks": total_checks,
        "high_duplicate": high_duplicate,
        "mid_duplicate": mid_duplicate,
        "low_duplicate": low_duplicate,
        "distribution": distribution,
        "daily_new": daily_new,
        "daily_checks": daily_checks,
        "top_checked": [dict(r) for r in top_checked]
    })

# ========= 查重 API =========

@app.route('/api/check', methods=['POST'])
@login_required
def api_check():
    data = request.get_json()
    customer_id = data.get('customer_id')
    compare_text = data.get('text', '').strip()

    if not customer_id or not compare_text:
        return jsonify({"error": "请选择客户并输入对比文本"}), 400

    db = get_db()
    customer = db.execute("SELECT * FROM customers WHERE id=?", (customer_id,)).fetchone()
    if not customer:
        return jsonify({"error": "客户不存在"}), 404

    # 全库查重：对比该客户的所有字段
    customer_text = ' '.join(filter(None, [
        customer['name'] or '',
        customer['phone'] or '',
        customer['company'] or '',
        customer['notes'] or '',
        customer['content'] or ''
    ]))

    result = check_duplicate(customer_text, compare_text)
    db.execute("INSERT INTO check_history (customer1_id, similarity, level) VALUES (?, ?, ?)",
               (customer_id, result['overall'], result['level']))
    today_date = bj_now().strftime("%Y-%m-%d")
    dup_count = 1 if result['overall'] >= 60 else 0
    db.execute("INSERT INTO check_logs (check_date, check_type, duplicate_count, total_checked) VALUES (?, ?, ?, ?)",
               (today_date, 'single', dup_count, 1))
    db.commit()

    result['customer_name'] = customer['name']
    result['customer_phone'] = customer['phone']
    result['customer_company'] = customer['company']
    return jsonify(result)

@app.route('/api/check-between', methods=['POST'])
@login_required
def api_check_between():
    data = request.get_json()
    id1 = data.get('customer_id_1')
    id2 = data.get('customer_id_2')

    if not id1 or not id2:
        return jsonify({"error": "请选择两个客户"}), 400
    if id1 == id2:
        return jsonify({"error": "不能和同一个客户对比"}), 400

    db = get_db()
    c1 = db.execute("SELECT * FROM customers WHERE id=?", (id1,)).fetchone()
    c2 = db.execute("SELECT * FROM customers WHERE id=?", (id2,)).fetchone()
    if not c1 or not c2:
        return jsonify({"error": "客户不存在"}), 404

    text1 = (c1['content'] or '') + ' ' + (c1['name'] or '') + ' ' + (c1['phone'] or '') + ' ' + (c1['company'] or '')
    text2 = (c2['content'] or '') + ' ' + (c2['name'] or '') + ' ' + (c2['phone'] or '') + ' ' + (c2['company'] or '')

    result = check_duplicate(text1, text2)
    db.execute("INSERT INTO check_history (customer1_id, customer2_id, similarity, level) VALUES (?, ?, ?, ?)",
               (id1, id2, result['overall'], result['level']))
    today_date = bj_now().strftime("%Y-%m-%d")
    dup_count = 1 if result['overall'] >= 60 else 0
    db.execute("INSERT INTO check_logs (check_date, check_type, duplicate_count, total_checked) VALUES (?, ?, ?, ?)",
               (today_date, 'between', dup_count, 1))
    db.commit()

    result['customer1'] = {'name': c1['name'], 'phone': c1['phone'], 'company': c1['company']}
    result['customer2'] = {'name': c2['name'], 'phone': c2['phone'], 'company': c2['company']}
    return jsonify(result)

@app.route('/api/batch-check', methods=['POST'])
@login_required
def api_batch_check():
    data = request.get_json()
    customer_id = data.get('customer_id')

    if not customer_id:
        return jsonify({"error": "请选择客户"}), 400

    db = get_db()
    target = db.execute("SELECT * FROM customers WHERE id=?", (customer_id,)).fetchone()
    if not target:
        return jsonify({"error": "客户不存在"}), 404

    all_others = db.execute("SELECT * FROM customers WHERE id != ?", (customer_id,)).fetchall()
    target_text = (target['content'] or '') + ' ' + (target['name'] or '') + ' ' + (target['phone'] or '')

    results = []
    for other in all_others:
        other_text = (other['content'] or '') + ' ' + (other['name'] or '') + ' ' + (other['phone'] or '')
        r = check_duplicate(target_text, other_text)
        results.append({
            'id': other['id'],
            'name': other['name'],
            'phone': other['phone'],
            'company': other['company'],
            'similarity': r['overall'],
            'level': r['level']
        })

    results.sort(key=lambda x: x['similarity'], reverse=True)
    today_date = bj_now().strftime("%Y-%m-%d")
    high_dup = sum(1 for r in results if r['similarity'] >= 60)
    db.execute("INSERT INTO check_logs (check_date, check_type, duplicate_count, total_checked) VALUES (?, ?, ?, ?)",
               (today_date, 'batch', high_dup, len(results)))
    db.commit()
    return jsonify({
        'target': {'id': target['id'], 'name': target['name'], 'phone': target['phone']},
        'results': [r for r in results if r['similarity'] > 10]
    })

@app.route('/api/quick-check', methods=['POST'])
@login_required
def api_quick_check():
    """快速查重：输入姓名或电话，全局搜索匹配客户"""
    data = request.get_json()
    keyword = data.get('keyword', '').strip()
    if not keyword:
        return jsonify({"error": "请输入姓名或电话号码"}), 400

    db = get_db()
    keyword_like = f'%{keyword}%'
    
    # 在姓名、电话、公司、内容中搜索匹配项
    rows = db.execute("""
        SELECT id, name, phone, email, company, notes, content, created_at
        FROM customers
        WHERE name LIKE ? OR phone LIKE ? OR company LIKE ? OR email LIKE ? OR content LIKE ?
        ORDER BY id DESC
    """, (keyword_like, keyword_like, keyword_like, keyword_like, keyword_like)).fetchall()
    
    results = []
    for r in rows:
        match_fields = []
        if keyword.lower() in (r['name'] or '').lower():
            match_fields.append('姓名')
        if keyword in (r['phone'] or ''):
            match_fields.append('电话')
        if keyword.lower() in (r['company'] or '').lower():
            match_fields.append('公司')
        if keyword.lower() in (r['email'] or '').lower():
            match_fields.append('邮箱')
        if keyword.lower() in (r['content'] or '').lower():
            match_fields.append('内容')
            
        results.append({
            'id': r['id'],
            'name': r['name'],
            'phone': r['phone'],
            'company': r['company'],
            'email': r['email'],
            'notes': r['notes'],
            'match_fields': match_fields,
            'created_at': r['created_at']
        })

    # Log to check_logs
    today_date = bj_now().strftime("%Y-%m-%d")
    dup_count = sum(1 for r in results if '姓名' in r.get('match_fields', []) or '电话' in r.get('match_fields', []))
    db.execute("INSERT INTO check_logs (check_date, check_type, duplicate_count, total_checked) VALUES (?, ?, ?, ?)",
               (today_date, 'quick', dup_count, len(results)))
    db.commit()

    return jsonify({
        'keyword': keyword,
        'total': len(results),
        'results': results,
        'checked_at': bj_now().strftime('%Y-%m-%d %H:%M:%S')
    })


@app.route('/api/quick-check-all', methods=['POST'])
@login_required
def api_quick_check_all():
    """批量快速查重：输入多个姓名/电话（每行一个），分别搜索匹配"""
    data = request.get_json()
    keywords = data.get('keywords', '').strip()
    if not keywords:
        return jsonify({"error": "请输入姓名或电话号码"}), 400

    lines = [k.strip() for k in keywords.split('\n') if k.strip()]
    if not lines:
        return jsonify({"error": "请输入至少一个关键词"}), 400

    db = get_db()
    all_results = []

    for keyword in lines:
        keyword_like = f'%{keyword}%'
        rows = db.execute("""
            SELECT id, name, phone, company, created_at
            FROM customers
            WHERE name LIKE ? OR phone LIKE ?
            ORDER BY id DESC
        """, (keyword_like, keyword_like)).fetchall()

        for r in rows:
            match_fields = []
            if keyword.lower() in (r['name'] or '').lower():
                match_fields.append('姓名')
            if keyword in (r['phone'] or ''):
                match_fields.append('电话')
            if keyword.lower() in (r['company'] or '').lower():
                match_fields.append('公司')

            all_results.append({
                'keyword': keyword,
                'id': r['id'],
                'name': r['name'],
                'phone': r['phone'],
                'company': r['company'],
                'match_fields': match_fields,
                'created_at': r['created_at'],
            })

    # Log to check_logs
    today_date = bj_now().strftime("%Y-%m-%d")
    db.execute("INSERT INTO check_logs (check_date, check_type, duplicate_count, total_checked) VALUES (?, ?, ?, ?)",
               (today_date, 'batch_quick', len(all_results), len(all_results)))
    db.commit()

    return jsonify({
        'keywords_count': len(lines),
        'total': len(all_results),
        'results': all_results,
        'checked_at': bj_now().strftime('%Y-%m-%d %H:%M:%S')
    })


@app.route('/api/history')
@login_required
def api_history():
    db = get_db()
    limit = min(int(request.args.get('limit', 50)), 200)
    rows = db.execute("""
        SELECT h.*, c1.name as c1_name, c1.phone as c1_phone, c2.name as c2_name, c2.phone as c2_phone
        FROM check_history h
        LEFT JOIN customers c1 ON h.customer1_id = c1.id
        LEFT JOIN customers c2 ON h.customer2_id = c2.id
        ORDER BY h.checked_at DESC LIMIT ?
    """, (limit,)).fetchall()
    return jsonify([dict(r) for r in rows])



# ========= 功能1: 批量删除重复客户 =========
@app.route('/api/customers/batch-clean-duplicates', methods=['GET'])
@login_required
def api_batch_clean_duplicates():
    db = get_db()
    name_groups = db.execute("""
        SELECT name, GROUP_CONCAT(id) as ids, COUNT(*) as cnt
        FROM customers WHERE name IS NOT NULL AND name != ''
        GROUP BY name HAVING cnt > 1
    """).fetchall()
    phone_groups = db.execute("""
        SELECT phone, GROUP_CONCAT(id) as ids, COUNT(*) as cnt
        FROM customers WHERE phone IS NOT NULL AND phone != ''
        GROUP BY phone HAVING cnt > 1
    """).fetchall()
    groups = []
    for g in name_groups:
        ids = [int(x) for x in g['ids'].split(',')]
        customers = db.execute(f"SELECT * FROM customers WHERE id IN ({','.join('?' * len(ids))})", ids).fetchall()
        groups.append({'type': 'name', 'key': g['name'], 'count': g['cnt'], 'customers': [dict(c) for c in customers]})
    for g in phone_groups:
        ids = [int(x) for x in g['ids'].split(',')]
        customers = db.execute(f"SELECT * FROM customers WHERE id IN ({','.join('?' * len(ids))})", ids).fetchall()
        groups.append({'type': 'phone', 'key': g['phone'], 'count': g['cnt'], 'customers': [dict(c) for c in customers]})
    return jsonify({'total_groups': len(groups), 'groups': groups})

@app.route('/api/customers/batch-clean-duplicates/action', methods=['POST'])
@login_required
def api_batch_clean_duplicates_action():
    data = request.get_json()
    action = data.get('action', '')
    customer_ids = data.get('ids', [])
    pwd = data.get('pwd', '')
    if pwd != DATA_PASSWORD:
        return jsonify({"error": "密码错误"}), 403
    if not customer_ids:
        return jsonify({"error": "请选择客户"}), 400
    db = get_db()
    if action == 'keep_one':
        delete_ids = customer_ids[1:]
        if delete_ids:
            placeholders = ','.join('?' * len(delete_ids))
            db.execute(f"DELETE FROM customers WHERE id IN ({placeholders})", delete_ids)
    elif action == 'delete_all':
        placeholders = ','.join('?' * len(customer_ids))
        db.execute(f"DELETE FROM customers WHERE id IN ({placeholders})", customer_ids)
    else:
        return jsonify({"error": "无效的操作类型"}), 400
    db.commit()
    return jsonify({"ok": True})

# ========= 功能3: 导入预览 =========
@app.route('/api/customers/import-preview', methods=['POST'])
@login_required
def api_import_preview():
    if 'file' not in request.files:
        return jsonify({"error": "请上传文件"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "文件不能为空"}), 400

    records = []
    filename = file.filename.lower()
    try:
        if filename.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            for row in reader:
                records.append(row)
        elif filename.endswith('.xlsx'):
            try:
                import openpyxl
            except ImportError:
                return jsonify({"error": "导入xlsx需要安装openpyxl"}), 500
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            header_row = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                record = {}
                for i, val in enumerate(row):
                    if i < len(header_row) and header_row[i]:
                        record[header_row[i]] = str(val) if val is not None else ''
                if record:
                    records.append(record)
        elif filename.endswith('.txt'):
            content = file.read().decode('utf-8-sig')
            lines = [l.strip() for l in content.split('\n') if l.strip()]
            for line in lines:
                record = {}
                parts = None
                for sep in ['\t', ',', '|', '，', ';', '；']:
                    tp = line.split(sep)
                    if len(tp) >= 2:
                        parts = tp
                        break
                if parts is None:
                    tp = re2.split(r'\s{2,}', line)
                    if len(tp) >= 2:
                        parts = tp
                if parts is None:
                    tp = line.split()
                    if len(tp) >= 2:
                        first = tp[0].strip()
                        if re.match(r'^[\+\d][\d\s\-\(\)]{4,}$', first.replace(' ', '')):
                            parts = tp
                        else:
                            for i, p in enumerate(tp):
                                p = p.strip()
                                if re.match(r'^[\+\d][\d\s\-\(\)]{6,}$', p.replace(' ', '')) and i > 0:
                                    parts = [' '.join(tp[:i]), p] + tp[i+1:]
                                    break
                            if parts is None:
                                parts = [line]
                if parts is None:
                    parts = [line]
                if len(parts) >= 2:
                    first = parts[0].strip()
                    second = parts[1].strip()
                    if re.match(r'^[\+\d][\d\s\-\(\)]{4,}$', first):
                        record['电话'] = first
                        record['姓名'] = second
                    else:
                        record['姓名'] = first
                        record['电话'] = second
                    if len(parts) >= 3:
                        record['公司'] = parts[2].strip()
                    if len(parts) >= 4:
                        record['备注'] = parts[3].strip()
                    records.append(record)
                elif len(parts) == 1 and line:
                    record['姓名'] = parts[0].strip()
                    records.append(record)
        else:
            return jsonify({"error": "仅支持 CSV、XLSX 和 TXT 格式"}), 400
    except Exception as e:
        return jsonify({"error": f"文件解析失败: {str(e)}"}), 400

    if not records:
        return jsonify({"error": "文件中没有数据"}), 400

    field_map = {
        '姓名': 'name', '名字': 'name', 'name': 'name',
        '电话': 'phone', '手机': 'phone', 'phone': 'phone', 'tel': 'phone',
        '邮箱': 'email', '邮件': 'email', 'email': 'email',
        '公司': 'company', '企业': 'company', '单位': 'company', 'company': 'company',
        '备注': 'notes', '备注': 'notes', 'notes': 'notes',
        '查重内容': 'content', '内容': 'content', 'content': 'content',
    }

    preview = []
    db = get_db()
    existing_names = set(r['name'] for r in db.execute("SELECT name FROM customers WHERE name IS NOT NULL").fetchall())
    existing_phones = set(r['phone'] for r in db.execute("SELECT phone FROM customers WHERE phone IS NOT NULL AND phone != ''").fetchall())

    duplicate_count = 0
    for idx, record in enumerate(records):
        mapped = {}
        for k, v in record.items():
            key = k.strip()
            if key in field_map:
                mapped[field_map[key]] = v.strip() if v else ''
        if 'name' not in mapped:
            for k in record:
                if any(term in k for term in ['姓名', '名字', '名称', 'name']):
                    mapped['name'] = record[k].strip() if record[k] else ''
                    break
        name = mapped.get('name', '')
        phone = clean_phone(mapped.get('phone', '')) if mapped.get('phone') else ''
        is_dup = (name and name in existing_names) or (phone and phone in existing_phones)
        if is_dup:
            duplicate_count += 1
        if idx < 20:
            preview.append({
                'index': idx + 1,
                'name': name,
                'phone': phone,
                'email': mapped.get('email', ''),
                'company': mapped.get('company', ''),
                'notes': mapped.get('notes', ''),
                'is_duplicate': is_dup
            })

    return jsonify({
        'ok': True,
        'total': len(records),
        'preview': preview,
        'duplicate_count': duplicate_count,
        'filename': file.filename
    })

# ========= 功能4: 归属地区域统计 =========
@app.route('/api/analytics/region-map', methods=['GET'])
@login_required
def api_region_map():
    db = get_db()
    rows = db.execute("""
        SELECT phone_region, COUNT(*) as count
        FROM customers
        WHERE phone_region IS NOT NULL AND phone_region != ''
        GROUP BY phone_region
        ORDER BY count DESC
    """).fetchall()
    regions = {r['phone_region']: r['count'] for r in rows}
    total_with = sum(regions.values())
    total_without = db.execute("SELECT COUNT(*) FROM customers WHERE phone_region IS NULL OR phone_region = ''").fetchone()[0]
    return jsonify({
        'regions': regions,
        'total_with_region': total_with,
        'total_without_region': total_without,
        'total_customers': total_with + total_without
    })

# ========= 功能6: 查重统计 =========
@app.route('/api/analytics/check-stats', methods=['GET'])
@login_required
def api_check_stats():
    db = get_db()
    today = bj_now().strftime("%Y-%m-%d")
    today_checks = db.execute("SELECT COALESCE(SUM(total_checked), 0) FROM check_logs WHERE check_date = ?", (today,)).fetchone()[0]
    total_checks = db.execute("SELECT COUNT(*) FROM check_logs").fetchone()[0]
    today_dup = db.execute("SELECT COALESCE(SUM(duplicate_count), 0) FROM check_logs WHERE check_date = ?", (today,)).fetchone()[0]
    total_dup = db.execute("SELECT COALESCE(SUM(duplicate_count), 0) FROM check_logs").fetchone()[0]
    avg_rate = round((total_dup / total_checks * 100) if total_checks > 0 else 0, 2)
    today_count = db.execute("SELECT COUNT(*) FROM check_logs WHERE check_date = ?", (today,)).fetchone()[0]
    weekly_trend = []
    for i in range(6, -1, -1):
        day = (bj_now() - timedelta(days=i)).strftime("%Y-%m-%d")
        count = db.execute("SELECT COALESCE(SUM(total_checked), 0) FROM check_logs WHERE check_date = ?", (day,)).fetchone()[0]
        weekly_trend.append({"date": day, "count": count})
    return jsonify({
        'today_checks': today_checks,
        'total_checks': total_checks,
        'today_duplicate': today_dup,
        'total_duplicate': total_dup,
        'avg_duplicate_rate': avg_rate,
        'today_check_count': today_count,
        'weekly_trend': weekly_trend
    })


# ========= 启动 =========
if __name__ == '__main__':
    print("=" * 60)
    print("  🚀 客户查重管理系统 (完整版)")
    print("=" * 60)
    print("  ✅ 生产模式启动: bash start.sh")
    print("  🌐 http://localhost:5000")
    print("  👤 admin / admin123")
    print("=" * 60)
    app.run(host='0.0.0.0', port=5000, debug=True)
