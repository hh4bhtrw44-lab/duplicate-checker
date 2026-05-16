#!/usr/bin/env python3
"""Apply all 6 backend features to app.py"""
import re

with open('app.py', 'r') as f:
    code = f.read()

# ===== Feature 6: Add check_logs table in init_db =====
old_table = '''        CREATE TABLE IF NOT EXISTS check_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer1_id INTEGER,
            customer2_id INTEGER,
            similarity REAL,
            level TEXT,
            checked_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (customer1_id) REFERENCES customers(id),
            FOREIGN KEY (customer2_id) REFERENCES customers(id)
        );'''
new_table = '''        CREATE TABLE IF NOT EXISTS check_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer1_id INTEGER,
            customer2_id INTEGER,
            similarity REAL,
            level TEXT,
            checked_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (customer1_id) REFERENCES customers(id),
            FOREIGN KEY (customer2_id) REFERENCES customers(id)
        );

        CREATE TABLE IF NOT EXISTS check_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            check_date TEXT,
            check_type TEXT,
            duplicate_count INTEGER DEFAULT 0,
            total_checked INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );'''
assert old_table in code, "check_history table not found!"
code = code.replace(old_table, new_table, 1)
print("✅ check_logs table added")

# ===== Feature 6: Add check logging in api_check =====
old_check1 = '''    result = check_duplicate(customer_text, compare_text)
    db.execute("INSERT INTO check_history (customer1_id, similarity, level) VALUES (?, ?, ?)",
               (customer_id, result['overall'], result['level']))
    db.commit()

    result['customer_name'] = customer['name']'''
new_check1 = '''    result = check_duplicate(customer_text, compare_text)
    db.execute("INSERT INTO check_history (customer1_id, similarity, level) VALUES (?, ?, ?)",
               (customer_id, result['overall'], result['level']))
    today_date = datetime.now().strftime("%Y-%m-%d")
    dup_count = 1 if result['overall'] >= 60 else 0
    db.execute("INSERT INTO check_logs (check_date, check_type, duplicate_count, total_checked) VALUES (?, ?, ?, ?)",
               (today_date, 'single', dup_count, 1))
    db.commit()

    result['customer_name'] = customer['name']'''
assert old_check1 in code, "api_check not found!"
code = code.replace(old_check1, new_check1, 1)
print("✅ api_check logging added")

# ===== Feature 6: Add check logging in api_check_between =====
old_check2 = '''    result = check_duplicate(text1, text2)
    db.execute("INSERT INTO check_history (customer1_id, customer2_id, similarity, level) VALUES (?, ?, ?, ?)",
               (id1, id2, result['overall'], result['level']))
    db.commit()

    result['customer1'] = {'name': c1['name']'''
new_check2 = '''    result = check_duplicate(text1, text2)
    db.execute("INSERT INTO check_history (customer1_id, customer2_id, similarity, level) VALUES (?, ?, ?, ?)",
               (id1, id2, result['overall'], result['level']))
    today_date = datetime.now().strftime("%Y-%m-%d")
    dup_count = 1 if result['overall'] >= 60 else 0
    db.execute("INSERT INTO check_logs (check_date, check_type, duplicate_count, total_checked) VALUES (?, ?, ?, ?)",
               (today_date, 'between', dup_count, 1))
    db.commit()

    result['customer1'] = {'name': c1['name']'''
assert old_check2 in code, "api_check_between not found!"
code = code.replace(old_check2, new_check2, 1)
print("✅ api_check_between logging added")

# ===== Feature 6: Add check logging in api_batch_check =====
old_check3 = '''    results.sort(key=lambda x: x['similarity'], reverse=True)
    return jsonify({
        'target': {'id': target['id'], 'name': target['name'], 'phone': target['phone']},
        'results': [r for r in results if r['similarity'] > 10]
    })'''
new_check3 = '''    results.sort(key=lambda x: x['similarity'], reverse=True)
    today_date = datetime.now().strftime("%Y-%m-%d")
    high_dup = sum(1 for r in results if r['similarity'] >= 60)
    db.execute("INSERT INTO check_logs (check_date, check_type, duplicate_count, total_checked) VALUES (?, ?, ?, ?)",
               (today_date, 'batch', high_dup, len(results)))
    db.commit()
    return jsonify({
        'target': {'id': target['id'], 'name': target['name'], 'phone': target['phone']},
        'results': [r for r in results if r['similarity'] > 10]
    })'''
assert old_check3 in code, "api_batch_check not found!"
code = code.replace(old_check3, new_check3, 1)
print("✅ api_batch_check logging added")

# ===== Feature 6: Add check logging in api_quick_check =====
old_quick = '''    return jsonify({
        'keyword': keyword,
        'total': len(results),
        'results': results,
        'checked_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })'''
new_quick = '''    today_date = datetime.now().strftime("%Y-%m-%d")
    dup_count = sum(1 for r in results if '姓名' in r.get('match_fields', []) or '电话' in r.get('match_fields', []))
    db.execute("INSERT INTO check_logs (check_date, check_type, duplicate_count, total_checked) VALUES (?, ?, ?, ?)",
               (today_date, 'quick', dup_count, len(results)))
    db.commit()
    return jsonify({
        'keyword': keyword,
        'total': len(results),
        'results': results,
        'checked_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })'''

# Find the specific return for quick_check (between results building and the blank line before quick_check_all)
idx = code.find("return jsonify({'keyword': keyword,'total': len(results),")
if idx > 0:
    # Get full match with proper whitespace
    old_block = '''    return jsonify({
        'keyword': keyword,
        'total': len(results),
        'results': results,
        'checked_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })'''
    new_block = '''    # Log to check_logs
    today_date = datetime.now().strftime("%Y-%m-%d")
    dup_count = sum(1 for r in results if '姓名' in r.get('match_fields', []) or '电话' in r.get('match_fields', []))
    db.execute("INSERT INTO check_logs (check_date, check_type, duplicate_count, total_checked) VALUES (?, ?, ?, ?)",
               (today_date, 'quick', dup_count, len(results)))
    db.commit()
    return jsonify({
        'keyword': keyword,
        'total': len(results),
        'results': results,
        'checked_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })'''
    
    # Find the quick_check return, not quick_check_all
    # The return at the first occurrence is from quick_check 
    first_occurrence = code.find("return jsonify({'keyword': keyword,'total': len(results),")
    if first_occurrence >= 0:
        code = code.replace(
            "return jsonify({'keyword': keyword,'total': len(results),",
            "# Log to check_logs\ntoday_date = datetime.now().strftime(\"%Y-%m-%d\")\ndup_count = sum(1 for r in results if '姓名' in r.get('match_fields', []) or '电话' in r.get('match_fields', []))\ndb.execute(\"INSERT INTO check_logs (check_date, check_type, duplicate_count, total_checked) VALUES (?, ?, ?, ?)\",\n           (today_date, 'quick', dup_count, len(results)))\ndb.commit()\nreturn jsonify({'keyword': keyword,'total': len(results),",
            1
        )
        print("✅ api_quick_check logging added")
    else:
        print("❌ api_quick_check return not found")
else:
    print("⚠️  Could not find quick_check return - checking manually")

# ===== Feature 6: Add check logging in api_quick_check_all =====
old_qall = '''            })
    return jsonify({
        'keywords_count': len(lines),
        'total': len(all_results),
        'results': all_results,
        'checked_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })'''
new_qall = '''            })
    today_date = datetime.now().strftime("%Y-%m-%d")
    db.execute("INSERT INTO check_logs (check_date, check_type, duplicate_count, total_checked) VALUES (?, ?, ?, ?)",
               (today_date, 'batch_quick', len(all_results), len(all_results)))
    db.commit()
    return jsonify({
        'keywords_count': len(lines),
        'total': len(all_results),
        'results': all_results,
        'checked_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })'''
assert old_qall in code, "api_quick_check_all not found!"
code = code.replace(old_qall, new_qall, 1)
print("✅ api_quick_check_all logging added")

# ===== Feature 2: Enhanced export with filters =====
old_export = '''    db = get_db()
    customers = db.execute("SELECT * FROM customers ORDER BY id DESC").fetchall()'''
new_export = '''    db = get_db()
    
    start_date = request.args.get('start_date', '').strip()
    end_date = request.args.get('end_date', '').strip()
    region = request.args.get('region', '').strip()
    search = request.args.get('search', '').strip()

    where_clauses = []
    params = []
    if start_date:
        where_clauses.append("DATE(created_at) >= ?")
        params.append(start_date)
    if end_date:
        where_clauses.append("DATE(created_at) <= ?")
        params.append(end_date)
    if region:
        where_clauses.append("phone_region = ?")
        params.append(region)
    if search:
        where_clauses.append("(name LIKE ? OR phone LIKE ? OR company LIKE ? OR email LIKE ?)")
        params.extend([f'%{search}%'] * 4)
    
    where_sql = ""
    if where_clauses:
        where_sql = "WHERE " + " AND ".join(where_clauses)
    
    customers = db.execute(f"SELECT * FROM customers {where_sql} ORDER BY id DESC", params).fetchall()'''
assert old_export in code, "export customers query not found!"
code = code.replace(old_export, new_export, 1)
print("✅ Export enhanced with filters")

# ===== Feature 5: Enhanced customer search with region/date/company =====
old_customers_get = '''@app.route('/api/customers', methods=['GET'])
@login_required
def api_customers():
    db = get_db()
    search = request.args.get('search', '').strip()
    search_field = request.args.get('field', 'all')  # 增强搜索
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))
    offset = (page - 1) * per_page

    # 限制最大每页条数
    if per_page > 100:
        per_page = 100

    if search:
        if search_field == 'name':
            condition = "name LIKE ?"
        elif search_field == 'phone':
            condition = "phone LIKE ?"
        elif search_field == 'email':
            condition = "email LIKE ?"
        elif search_field == 'company':
            condition = "company LIKE ?"
        else:
            condition = "name LIKE ? OR phone LIKE ? OR email LIKE ? OR company LIKE ? OR notes LIKE ?"

        if search_field == 'all':
            rows = db.execute(
                f"SELECT * FROM customers WHERE {condition} ORDER BY id DESC LIMIT ? OFFSET ?",
                (f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%', per_page, offset)
            ).fetchall()
            total = db.execute(
                f"SELECT COUNT(*) FROM customers WHERE {condition}",
                (f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%')
            ).fetchone()[0]
        else:
            rows = db.execute(
                f"SELECT * FROM customers WHERE {condition} ORDER BY id DESC LIMIT ? OFFSET ?",
                (f'%{search}%', per_page, offset)
            ).fetchall()
            total = db.execute(
                f"SELECT COUNT(*) FROM customers WHERE {condition}",
                (f'%{search}%',)
            ).fetchone()[0]
    else:
        rows = db.execute("SELECT * FROM customers ORDER BY id DESC LIMIT ? OFFSET ?", (per_page, offset)).fetchall()
        total = db.execute("SELECT COUNT(*) FROM customers").fetchone()[0]

    return jsonify({
        "customers": [dict(r) for r in rows],
        "total": total,
        "page": page,
        "per_page": per_page
    })'''
new_customers_get = '''@app.route('/api/customers', methods=['GET'])
@login_required
def api_customers():
    db = get_db()
    search = request.args.get('search', '').strip()
    search_field = request.args.get('field', 'all')
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))
    offset = (page - 1) * per_page

    # 高级搜索参数
    region_filter = request.args.get('region', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    company_filter = request.args.get('company', '').strip()

    if per_page > 100:
        per_page = 100

    where_clauses = []
    params = []

    if search:
        if search_field == 'name':
            where_clauses.append("name LIKE ?")
            params.append(f'%{search}%')
        elif search_field == 'phone':
            where_clauses.append("phone LIKE ?")
            params.append(f'%{search}%')
        elif search_field == 'email':
            where_clauses.append("email LIKE ?")
            params.append(f'%{search}%')
        elif search_field == 'company':
            where_clauses.append("company LIKE ?")
            params.append(f'%{search}%')
        else:
            where_clauses.append("(name LIKE ? OR phone LIKE ? OR email LIKE ? OR company LIKE ? OR notes LIKE ?)")
            params.extend([f'%{search}%'] * 5)
    
    if region_filter:
        where_clauses.append("phone_region = ?")
        params.append(region_filter)
    if date_from:
        where_clauses.append("DATE(created_at) >= ?")
        params.append(date_from)
    if date_to:
        where_clauses.append("DATE(created_at) <= ?")
        params.append(date_to)
    if company_filter:
        where_clauses.append("company LIKE ?")
        params.append(f'%{company_filter}%')

    where_sql = ""
    if where_clauses:
        where_sql = "WHERE " + " AND ".join(where_clauses)

    rows = db.execute(f"SELECT * FROM customers {where_sql} ORDER BY id DESC LIMIT ? OFFSET ?",
                      params + [per_page, offset]).fetchall()
    total = db.execute(f"SELECT COUNT(*) FROM customers {where_sql}", params).fetchone()[0]

    return jsonify({
        "customers": [dict(r) for r in rows],
        "total": total,
        "page": page,
        "per_page": per_page
    })'''
assert old_customers_get in code, "api_customers GET not found!"
code = code.replace(old_customers_get, new_customers_get, 1)
print("✅ Advanced search (region/date/company) added to /api/customers")

# ===== Add new routes before __main__ =====
new_routes = '''
# ========= 功能1: 批量删除重复客户 =========
@app.route('/api/customers/batch-clean-duplicates', methods=['GET'])
@login_required
def api_batch_clean_duplicates():
    db = get_db()
    name_groups = db.execute("""
        SELECT name, GROUP_CONCAT(id) as ids, COUNT(*) as cnt
        FROM customers WHERE name IS NOT NULL AND name != ''
        GROUP BY name HAVING cnt > 1
    """).fetchall()
    phone_groups = db.execute("""
        SELECT phone, GROUP_CONCAT(id) as ids, COUNT(*) as cnt
        FROM customers WHERE phone IS NOT NULL AND phone != ''
        GROUP BY phone HAVING cnt > 1
    """).fetchall()
    groups = []
    for g in name_groups:
        ids = [int(x) for x in g['ids'].split(',')]
        customers = db.execute(f"SELECT * FROM customers WHERE id IN ({','.join('?' * len(ids))})", ids).fetchall()
        groups.append({'type': 'name', 'key': g['name'], 'count': g['cnt'], 'customers': [dict(c) for c in customers]})
    for g in phone_groups:
        ids = [int(x) for x in g['ids'].split(',')]
        customers = db.execute(f"SELECT * FROM customers WHERE id IN ({','.join('?' * len(ids))})", ids).fetchall()
        groups.append({'type': 'phone', 'key': g['phone'], 'count': g['cnt'], 'customers': [dict(c) for c in customers]})
    return jsonify({'total_groups': len(groups), 'groups': groups})

@app.route('/api/customers/batch-clean-duplicates/action', methods=['POST'])
@login_required
def api_batch_clean_duplicates_action():
    data = request.get_json()
    action = data.get('action', '')
    customer_ids = data.get('ids', [])
    pwd = data.get('pwd', '')
    if pwd != DATA_PASSWORD:
        return jsonify({"error": "密码错误"}), 403
    if not customer_ids:
        return jsonify({"error": "请选择客户"}), 400
    db = get_db()
    if action == 'keep_one':
        delete_ids = customer_ids[1:]
        if delete_ids:
            placeholders = ','.join('?' * len(delete_ids))
            db.execute(f"DELETE FROM customers WHERE id IN ({placeholders})", delete_ids)
    elif action == 'delete_all':
        placeholders = ','.join('?' * len(customer_ids))
        db.execute(f"DELETE FROM customers WHERE id IN ({placeholders})", customer_ids)
    else:
        return jsonify({"error": "无效的操作类型"}), 400
    db.commit()
    return jsonify({"ok": True})

# ========= 功能3: 导入预览 =========
@app.route('/api/customers/import-preview', methods=['POST'])
@login_required
def api_import_preview():
    if 'file' not in request.files:
        return jsonify({"error": "请上传文件"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "文件不能为空"}), 400

    records = []
    filename = file.filename.lower()
    try:
        if filename.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            for row in reader:
                records.append(row)
        elif filename.endswith('.xlsx'):
            try:
                import openpyxl
            except ImportError:
                return jsonify({"error": "导入xlsx需要安装openpyxl"}), 500
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            header_row = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                record = {}
                for i, val in enumerate(row):
                    if i < len(header_row) and header_row[i]:
                        record[header_row[i]] = str(val) if val is not None else ''
                if record:
                    records.append(record)
        elif filename.endswith('.txt'):
            content = file.read().decode('utf-8-sig')
            lines = [l.strip() for l in content.split('\\n') if l.strip()]
            for line in lines:
                record = {}
                parts = None
                for sep in ['\\t', ',', '|', '，', ';', '；']:
                    test_parts = line.split(sep)
                    if len(test_parts) >= 2:
                        parts = test_parts
                        break
                if parts is None:
                    import re as re2
                    test_parts = re2.split(r'\\s{2,}', line)
                    if len(test_parts) >= 2:
                        parts = test_parts
                if parts is None:
                    test_parts = line.split()
                    if len(test_parts) >= 2:
                        first = test_parts[0].strip()
                        if re.match(r'^[\\+\\d][\\d\\s\\-\\(\\)]{4,}$', first.replace(' ', '')):
                            parts = test_parts
                        else:
                            for i, p in enumerate(test_parts):
                                p = p.strip()
                                if re.match(r'^[\\+\\d][\\d\\s\\-\\(\\)]{6,}$', p.replace(' ', '')) and i > 0:
                                    parts = [' '.join(test_parts[:i]), p] + test_parts[i+1:]
                                    break
                            if parts is None:
                                parts = [line]
                if parts is None:
                    parts = [line]
                if len(parts) >= 2:
                    first = parts[0].strip()
                    second = parts[1].strip()
                    if re.match(r'^[\\+\\d][\\d\\s\\-\\(\\)]{4,}$', first):
                        record['电话'] = first
                        record['姓名'] = second
                    else:
                        record['姓名'] = first
                        record['电话'] = second
                    if len(parts) >= 3:
                        record['公司'] = parts[2].strip()
                    if len(parts) >= 4:
                        record['备注'] = parts[3].strip()
                    records.append(record)
                elif len(parts) == 1 and line:
                    record['姓名'] = parts[0].strip()
                    records.append(record)
        else:
            return jsonify({"error": "仅支持 CSV、XLSX 和 TXT 格式"}), 400
    except Exception as e:
        return jsonify({"error": f"文件解析失败: {str(e)}"}), 400

    if not records:
        return jsonify({"error": "文件中没有数据"}), 400

    field_map = {
        '姓名': 'name', '名字': 'name', 'name': 'name',
        '电话': 'phone', '手机': 'phone', 'phone': 'phone', 'tel': 'phone',
        '邮箱': 'email', '邮件': 'email', 'email': 'email',
        '公司': 'company', '企业': 'company', '单位': 'company', 'company': 'company',
        '备注': 'notes', '备注': 'notes', 'notes': 'notes',
        '查重内容': 'content', '内容': 'content', 'content': 'content',
    }

    preview = []
    db = get_db()
    existing_names = set(r['name'] for r in db.execute("SELECT name FROM customers WHERE name IS NOT NULL").fetchall())
    existing_phones = set(r['phone'] for r in db.execute("SELECT phone FROM customers WHERE phone IS NOT NULL AND phone != ''").fetchall())

    duplicate_count = 0
    for idx, record in enumerate(records):
        mapped = {}
        for k, v in record.items():
            key = k.strip()
            if key in field_map:
                mapped[field_map[key]] = v.strip() if v else ''
        if 'name' not in mapped:
            for k in record:
                if any(term in k for term in ['姓名', '名字', '名称', 'name']):
                    mapped['name'] = record[k].strip() if record[k] else ''
                    break
        name = mapped.get('name', '')
        phone = ''
        if mapped.get('phone'):
            phone = re.sub(r'[\\s\\+\\-\\(\\)]', '', mapped['phone'])
        is_dup = (name and name in existing_names) or (phone and phone in existing_phones)
        if is_dup:
            duplicate_count += 1
        if idx < 20:
            preview.append({
                'index': idx + 1,
                'name': name,
                'phone': phone,
                'email': mapped.get('email', ''),
                'company': mapped.get('company', ''),
                'notes': mapped.get('notes', ''),
                'is_duplicate': is_dup
            })

    return jsonify({
        'ok': True,
        'total': len(records),
        'preview': preview,
        'duplicate_count': duplicate_count,
        'filename': file.filename
    })


# ========= 功能4: 归属地区域统计 =========
@app.route('/api/analytics/region-map', methods=['GET'])
@login_required
def api_region_map():
    db = get_db()
    rows = db.execute("""
        SELECT phone_region, COUNT(*) as count
        FROM customers
        WHERE phone_region IS NOT NULL AND phone_region != ''
        GROUP BY phone_region
        ORDER BY count DESC
    """).fetchall()
    regions = {r['phone_region']: r['count'] for r in rows}
    total_with = sum(regions.values())
    total_without = db.execute("SELECT COUNT(*) FROM customers WHERE phone_region IS NULL OR phone_region = ''").fetchone()[0]
    return jsonify({
        'regions': regions,
        'total_with_region': total_with,
        'total_without_region': total_without,
        'total_customers': total_with + total_without
    })


# ========= 功能6: 查重统计 =========
@app.route('/api/analytics/check-stats', methods=['GET'])
@login_required
def api_check_stats():
    db = get_db()
    today = datetime.now().strftime("%Y-%m-%d")
    today_checks = db.execute("SELECT COALESCE(SUM(total_checked), 0) FROM check_logs WHERE check_date = ?", (today,)).fetchone()[0]
    total_checks = db.execute("SELECT COUNT(*) FROM check_logs").fetchone()[0]
    today_dup = db.execute("SELECT COALESCE(SUM(duplicate_count), 0) FROM check_logs WHERE check_date = ?", (today,)).fetchone()[0]
    total_dup = db.execute("SELECT COALESCE(SUM(duplicate_count), 0) FROM check_logs").fetchone()[0]
    avg_rate = round((total_dup / total_checks * 100) if total_checks > 0 else 0, 2)
    today_count = db.execute("SELECT COUNT(*) FROM check_logs WHERE check_date = ?", (today,)).fetchone()[0]
    weekly_trend = []
    for i in range(6, -1, -1):
        day = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
        count = db.execute("SELECT COALESCE(SUM(total_checked), 0) FROM check_logs WHERE check_date = ?", (day,)).fetchone()[0]
        weekly_trend.append({"date": day, "count": count})
    return jsonify({
        'today_checks': today_checks,
        'total_checks': total_checks,
        'today_duplicate': today_dup,
        'total_duplicate': total_dup,
        'avg_duplicate_rate': avg_rate,
        'today_check_count': today_count,
        'weekly_trend': weekly_trend
    })
'''

# Insert new routes before __main__
old_startup = '\n\n# ========= 启动 =========\nif __name__ == \'__main__\':'
assert old_startup in code, "__main__ not found!"
code = code.replace(old_startup, new_routes + old_startup, 1)
print("✅ New feature routes added (batch-clean, import-preview, region-map, check-stats)")

# Verify
try:
    compile(code, 'app.py', 'exec')
    print("✅ app.py compiles successfully!")
except SyntaxError as e:
    print(f"❌ SyntaxError: {e}")

with open('app.py', 'w') as f:
    f.write(code)

print("\nAll backend changes applied!")
